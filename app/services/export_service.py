"""Excel and PDF export service."""

from __future__ import annotations
from pathlib import Path
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.database.engine import db_session
from app.database.models.member import Member
from app.database.models.membership_fee import MembershipFeePeriod, MembershipFeePayment
from app.database.models.target_fee import TargetFeeCampaign, TargetFeePayment
from app.database.models.electricity import ElectricityPayment, MeterReading
from app.constants import MemberStatus


def export_members_excel(path: str | Path) -> None:
    """Export members list to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Участники"

    headers = ["ID", "Участок", "Фамилия", "Имя", "Отчество",
               "Площадь", "Телефон", "Email", "Статус"]
    _write_headers(ws, headers)

    with db_session(readonly=True) as session:
        members = session.query(Member).order_by(Member.plot_number).all()
        for i, m in enumerate(members, start=2):
            ws.cell(i, 1, m.id)
            ws.cell(i, 2, m.plot_number)
            ws.cell(i, 3, m.last_name)
            ws.cell(i, 4, m.first_name)
            ws.cell(i, 5, m.patronymic or "")
            ws.cell(i, 6, str(m.plot_area))
            ws.cell(i, 7, m.phone or "")
            ws.cell(i, 8, m.email or "")
            ws.cell(i, 9, m.status)

    _auto_width(ws)
    wb.save(path)


def export_membership_fees_excel(path: str | Path, period_id: int) -> None:
    """Export membership fee payments for a period."""
    wb = Workbook()
    ws = wb.active

    with db_session(readonly=True) as session:
        period = session.get(MembershipFeePeriod, period_id)
        ws.title = f"Взносы {period.year}" if period else "Взносы"

        headers = ["Участок", "ФИО", "Площадь", "Начислено", "Оплачено", "Долг", "Статус"]
        _write_headers(ws, headers)

        results = session.query(MembershipFeePayment, Member).join(
            Member, MembershipFeePayment.member_id == Member.id
        ).filter(
            MembershipFeePayment.period_id == period_id
        ).all()
        for i, (p, member) in enumerate(results, start=2):
            debt = p.amount_due - p.amount_paid
            ws.cell(i, 1, member.plot_number)
            ws.cell(i, 2, member.full_name)
            ws.cell(i, 3, str(member.plot_area))
            ws.cell(i, 4, str(p.amount_due))
            ws.cell(i, 5, str(p.amount_paid))
            ws.cell(i, 6, str(debt))
            ws.cell(i, 7, p.status)

    _auto_width(ws)
    wb.save(path)


def _setup_pdf_fonts(pdf) -> str:
    """Configure Cyrillic fonts: DejaVu -> Arial (Windows) -> Times New Roman -> Helvetica fallback.
    Returns the font family name to use.
    """
    import os
    from app.logger import get_logger
    _logger = get_logger(__name__)

    try:
        pdf.add_font("DejaVu", "", "DejaVuSans.ttf", uni=True)
        pdf.add_font("DejaVu", "B", "DejaVuSans-Bold.ttf", uni=True)
        return "DejaVu"
    except Exception:
        pass

    font_dir = os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts")

    try:
        pdf.add_font("Arial", "", os.path.join(font_dir, "arial.ttf"), uni=True)
        pdf.add_font("Arial", "B", os.path.join(font_dir, "arialbd.ttf"), uni=True)
        return "Arial"
    except Exception:
        pass

    try:
        pdf.add_font("TimesNR", "", os.path.join(font_dir, "times.ttf"), uni=True)
        pdf.add_font("TimesNR", "B", os.path.join(font_dir, "timesbd.ttf"), uni=True)
        return "TimesNR"
    except Exception:
        pass

    _logger.warning("No Cyrillic-capable font found — PDF may render incorrectly")
    return "Helvetica"


def export_members_pdf(path: str | Path) -> None:
    """Export members list to PDF with Cyrillic support."""
    from fpdf import FPDF

    pdf = FPDF()
    pdf.add_page(orientation="L")
    font_name = _setup_pdf_fonts(pdf)

    pdf.set_font(font_name, "B", 14)
    pdf.cell(0, 10, "Список участников СНТ", ln=True, align="C")
    pdf.set_font(font_name, "", 8)
    pdf.cell(0, 6, f"Дата: {date.today()}", ln=True, align="R")
    pdf.ln(5)

    headers = ["Участок", "ФИО", "Площадь", "Телефон", "Email", "Статус"]
    col_widths = [25, 80, 25, 40, 60, 30]

    pdf.set_font(font_name, "B", 9)
    for h, w in zip(headers, col_widths):
        pdf.cell(w, 8, h, border=1, align="C")
    pdf.ln()

    with db_session(readonly=True) as session:
        members = session.query(Member).order_by(Member.plot_number).all()
        pdf.set_font(font_name, "", 9)
        for m in members:
            status = "Акт." if m.status == MemberStatus.ACTIVE.value else "Арх."
            pdf.cell(col_widths[0], 7, m.plot_number, border=1)
            pdf.cell(col_widths[1], 7, m.full_name[:40], border=1)
            pdf.cell(col_widths[2], 7, str(m.plot_area), border=1, align="C")
            pdf.cell(col_widths[3], 7, m.phone or "", border=1)
            pdf.cell(col_widths[4], 7, m.email or "", border=1)
            pdf.cell(col_widths[5], 7, status, border=1, align="C")
            pdf.ln()

    pdf.output(str(path))


def export_debtors_pdf(path: str | Path) -> None:
    """Export debtors report to PDF."""
    from fpdf import FPDF
    from app.services.report_service import get_debtors_list

    pdf = FPDF()
    pdf.add_page()
    font_name = _setup_pdf_fonts(pdf)

    pdf.set_font(font_name, "B", 14)
    pdf.cell(0, 10, "Отчёт по должникам", ln=True, align="C")
    pdf.set_font(font_name, "", 8)
    pdf.cell(0, 6, f"Дата: {date.today()}", ln=True, align="R")
    pdf.ln(5)

    headers = ["Участок", "ФИО", "Общий долг (руб.)"]
    col_widths = [30, 90, 50]

    pdf.set_font(font_name, "B", 10)
    for h, w in zip(headers, col_widths):
        pdf.cell(w, 8, h, border=1, align="C")
    pdf.ln()

    debtors = get_debtors_list()
    pdf.set_font(font_name, "", 10)
    total = Decimal("0")
    for d in debtors:
        pdf.cell(col_widths[0], 7, d["plot_number"], border=1)
        pdf.cell(col_widths[1], 7, d["full_name"][:45], border=1)
        pdf.cell(col_widths[2], 7, f"{d['total_debt']:.2f}", border=1, align="R")
        pdf.ln()
        total += d["total_debt"]

    pdf.set_font(font_name, "B", 10)
    pdf.cell(col_widths[0] + col_widths[1], 8, "ИТОГО:", border=1, align="R")
    pdf.cell(col_widths[2], 8, f"{total:.2f}", border=1, align="R")

    pdf.output(str(path))


def _write_headers(ws, headers: list[str]):
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

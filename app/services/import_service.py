"""Excel/CSV import service."""

from __future__ import annotations
import csv
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from app.database.engine import db_session
from app.database.models.member import Member
from app.constants import MemberStatus
from app.services.audit_service import log_action
from app.constants import AuditAction


def import_members_csv(path: str | Path) -> dict:
    """Import members from CSV file.
    Expected columns: plot_number, last_name, first_name, patronymic, plot_area, phone, email

    Returns dict with counts: created, skipped, errors.
    """
    result = {"created": 0, "skipped": 0, "errors": []}

    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=";")
        if not reader.fieldnames:
            # Try comma delimiter
            f.seek(0)
            reader = csv.DictReader(f, delimiter=",")

        rows = list(reader)

    try:
        with db_session() as session:
            for i, row in enumerate(rows, start=2):
                try:
                    plot_number = _get_field(row, ["plot_number", "участок", "номер_участка", "Участок"])
                    last_name = _get_field(row, ["last_name", "фамилия", "Фамилия"])
                    first_name = _get_field(row, ["first_name", "имя", "Имя"])

                    if not plot_number or not last_name or not first_name:
                        result["errors"].append(f"Строка {i}: не хватает обязательных полей")
                        continue

                    # Skip if exists
                    existing = session.query(Member).filter(
                        Member.plot_number == plot_number
                    ).first()
                    if existing:
                        result["skipped"] += 1
                        continue

                    patronymic = _get_field(row, ["patronymic", "отчество", "Отчество"])
                    plot_area_str = _get_field(row, ["plot_area", "площадь", "Площадь"])
                    try:
                        plot_area = Decimal(plot_area_str) if plot_area_str else Decimal("6")
                    except (InvalidOperation, ValueError):
                        plot_area = Decimal("6")

                    phone = _get_field(row, ["phone", "телефон", "Телефон"])
                    email = _get_field(row, ["email", "Email", "EMAIL"])

                    member = Member(
                        plot_number=plot_number,
                        last_name=last_name,
                        first_name=first_name,
                        patronymic=patronymic or None,
                        plot_area=plot_area,
                        phone=phone or None,
                        email=email or None,
                        status=MemberStatus.ACTIVE.value,
                    )
                    session.add(member)
                    result["created"] += 1

                except Exception as e:
                    result["errors"].append(f"Строка {i}: {e}")

            if result["created"]:
                log_action(AuditAction.IMPORT.value, "member", None,
                           f"Импорт CSV: {result['created']} создано")
    except Exception as e:
        result["errors"].append(str(e))

    return result


def import_members_excel(path: str | Path) -> dict:
    """Import members from Excel file.
    First row should be headers.
    """
    result = {"created": 0, "skipped": 0, "errors": []}

    wb = load_workbook(path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return result

    # First row as headers
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    try:
        with db_session() as session:
            for i, row_values in enumerate(rows[1:], start=2):
                try:
                    row_dict = dict(zip(headers, row_values))

                    plot_number = str(
                        _get_field(row_dict, ["plot_number", "участок", "номер_участка", "id"]) or ""
                    ).strip()
                    last_name = str(
                        _get_field(row_dict, ["last_name", "фамилия"]) or ""
                    ).strip()
                    first_name = str(
                        _get_field(row_dict, ["first_name", "имя"]) or ""
                    ).strip()

                    if not plot_number or not last_name or not first_name:
                        result["errors"].append(f"Строка {i}: не хватает обязательных полей")
                        continue

                    existing = session.query(Member).filter(
                        Member.plot_number == plot_number
                    ).first()
                    if existing:
                        result["skipped"] += 1
                        continue

                    patronymic = _get_field(row_dict, ["patronymic", "отчество"])
                    plot_area_str = _get_field(row_dict, ["plot_area", "площадь"])
                    try:
                        plot_area = Decimal(str(plot_area_str)) if plot_area_str else Decimal("6")
                    except (InvalidOperation, ValueError):
                        plot_area = Decimal("6")

                    phone = _get_field(row_dict, ["phone", "телефон"])
                    email = _get_field(row_dict, ["email"])

                    member = Member(
                        plot_number=str(plot_number),
                        last_name=str(last_name),
                        first_name=str(first_name),
                        patronymic=str(patronymic) if patronymic else None,
                        plot_area=plot_area,
                        phone=str(phone) if phone else None,
                        email=str(email) if email else None,
                        status=MemberStatus.ACTIVE.value,
                    )
                    session.add(member)
                    result["created"] += 1

                except Exception as e:
                    result["errors"].append(f"Строка {i}: {e}")

            if result["created"]:
                log_action(AuditAction.IMPORT.value, "member", None,
                           f"Импорт Excel: {result['created']} создано")
    except Exception as e:
        result["errors"].append(str(e))

    wb.close()
    return result


def _get_field(row: dict, possible_keys: list[str]) -> str | None:
    """Try multiple possible column names (case-insensitive)."""
    for key in possible_keys:
        # Exact match first
        if key in row and row[key]:
            return str(row[key]).strip()
        # Case-insensitive
        for k in row:
            if k and k.strip().lower() == key.lower() and row[k]:
                return str(row[k]).strip()
    return None
